#!/usr/bin/env python
# encoding: utf-8
'''
@author: zhanghui
@file: excel_sum_v0.3.py
@time: 2018/9/10 19:12
'''

import tkinter as tk
from tkinter import messagebox as msg
import tkinter.filedialog as tkFileDialog
import time
from getExcel import getExcel as excel
from openpyxl import load_workbook
class excelsum(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("总结tool v0.3 \t by 张珲")
        self.geometry("300x200")
        self.file_opt = {}
        self.file_list = []
        self.file_opt['initialfile'] = ''
        self.file_opt['filetypes'] = [('xls', 'xls'), ('xlsx', 'xlsx')]
        self.file_opt['initialdir'] = '.'
        self.today = time.strftime("%Y/%m/%d")
        self.wm_attributes('-topmost', 1)
        self.resizable(width=False, height=False)
        # borders = xlwt.Borders()
        # borders.left = 1
        # borders.right = 1
        # borders.top = 1
        # borders.bottom = 1
        # borders.bottom_colour = 0x3A
        # alignment = xlwt.Alignment()
        # alignment.horz = xlwt.Alignment.HORZ_CENTER
        # alignment.vert = xlwt.Alignment.VERT_CENTER
        # font = xlwt.Font()
        # font.name = 'SimSun'  # 指定“宋体”
        # self.style1 = xlwt.XFStyle()
        # self.style1.num_format_str = 'YYYY/M/D'
        # self.style2 = xlwt.XFStyle()
        # self.style2.borders = borders
        # self.style2.font = font
        # self.style1.borders = borders
        # self.style1.alignment = alignment
        # self.style2.alignment = alignment
        self.result = []
        self.file_count = 0
        self.row_flag = 0
        self.Rightframe = tk.Frame(self,height=200,width=150)
        self.Leftframe = tk.Frame(self, height=200, width=150)
        self.L_up_frame = tk.Frame(self.Leftframe, height=50, width=150)
        self.R_m_frame = tk.Frame(self.Rightframe, height=100, width=150)
        self.init_frame()
        self.open_botton = tk.Button(self.L_up_frame, height=1, width=6,text="打开", fg="blue", bg="white", command=self.askopenfile)
        self.open_botton.pack()
        self.show_label = tk.Label(self.L_up_frame, text="已载入 %s 个" % self.file_count,justify='left')
        self.show_label.pack()
        self.input_text = tk.Entry(self.R_m_frame,width=15, font=('Courier New', 10))
        self.input_text.insert(0, self.today)
        self.input_text.pack()
        self.start_botton = tk.Button(self.R_m_frame, height=1, width=6, text="开始", fg="blue", bg="white",
                                     command=self.start)
        self.start_botton.pack()
    def init_frame(self):
        self.Rightframe.pack(side=tk.RIGHT);self.Leftframe.pack(side=tk.LEFT)
        self.L_up_frame.pack(side=tk.TOP);#self.L_down_frame.pack(side=tk.BOTTOM)
        self.R_m_frame.pack(fill=tk.Y)
    def askopenfile(self):
        self.file_opt['multiple'] = True
        self.file_opt['filetypes'] = [('xls', 'xls'), ('xlsx', 'xlsx')]
        filename = tkFileDialog.askopenfilename(**self.file_opt)
        if filename:
            self.file_count += len(filename)
            for files in filename:
                if files not in self.file_list:
                    self.show_label.config(text="已载入 %s 个" % self.file_count)
                    self.file_list.append(files)
    def start(self):
        text = self.input_text.get()
        if text:
            import re
            if re.match('[0-9]{4}/[0-9]{1,2}/[0-9]{1,2}',text):
                self.now = text
                self.startsum()
            else:
                msg.showinfo("提示！", "日期格式有误，如 2018/1/1")
        else:
            msg.showinfo("提示！", "请输入待提取日期，如 2018/1/1")
    def datafield(self,afile):
        demo = excel(afile)
        demo.new = self.now
        for index in demo.get():
            yield index[1]
    def startsum(self):
        if self.file_list:
            for excel in self.file_list:
                for rowdata in self.datafield(excel):
                    self.result.append(rowdata)
            self.opensavafile()
        else:
            msg.showinfo("提示！", "尚未载入任何文件")
    def opensavafile(self):
        self.file_opt['multiple'] = None
        self.file_opt['filetypes'] = [('xlsx', 'xlsx')]
        filename = tkFileDialog.askopenfilename(**self.file_opt)
        if filename:
            workbook = load_workbook(filename)
            worksheet = workbook['Sheet1']
            style = worksheet.row_dimensions[3]
            row = worksheet.max_row + 1
            for data in self.result:
                for index in range(len(data)):
                    worksheet['%d' % row][index].value = data[index]
                row += 1
            workbook.save(filename)
        msg.showinfo("提示！", "%s 共添加 %s 条" %(self.now,len(self.result)))


if __name__ == "__main__":
    demo1 = excelsum()
    demo1.mainloop()
